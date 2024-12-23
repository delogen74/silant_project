import datetime
import openpyxl
from django.core.management.base import BaseCommand
from machines.models import Machine
from maintenance.models import TO
from claims.models import Reclamation


class Command(BaseCommand):
    help = "Импорт данных из Excel-файлов в модели Machine, TO, Reclamation"

    def handle(self, *args, **options):
        self.import_machines('machines.xlsx')
        self.import_to('to_output.xlsx')
        self.import_claims('reklamacia_output.xlsx')

    # ------------------ 1. Импорт Машин ------------------------
    def import_machines(self, filename):
        self.stdout.write(f"Импорт машин из {filename}...")
        try:
            wb = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f"Файл {filename} не найден!"))
            return

        sheet = wb.active
        row_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):


            if len(row) < 17:
                self.stdout.write("Строка короче 17 столбцов, пропускаем: " + str(row))
                continue


            model_tech = row[1] or ''
            factory_num = row[2] or ''
            engine_mod = row[3] or ''
            engine_num = row[4] or ''
            transm_mod = row[5] or ''
            transm_num = row[6] or ''
            drive_axle_mod = row[7] or ''
            drive_axle_num = row[8] or ''
            steer_axle_mod = row[9] or ''
            steer_axle_num = row[10] or ''
            shipment_date = row[11]
            buyer = row[12] or ''
            consignee = row[13] or ''
            delivery_addr = row[14] or ''
            equipment = row[15] or ''
            service_comp = row[16] or ''

            if not factory_num:
                self.stdout.write("Пустой factory_number, пропускаем строку.")
                continue

            shipment_date = self._parse_date(shipment_date)

            obj, created = Machine.objects.update_or_create(
                factory_number=factory_num,
                defaults={
                    'machine_model': model_tech,
                    'engine_model': engine_mod,
                    'engine_number': engine_num,
                    'transmission_model': transm_mod,
                    'transmission_number': transm_num,
                    'drive_axle_model': drive_axle_mod,
                    'drive_axle_number': drive_axle_num,
                    'steerable_axle_model': steer_axle_mod,
                    'steerable_axle_number': steer_axle_num,
                    'shipment_date': shipment_date,
                    'client': buyer,
                    'consignee': consignee,
                    'delivery_address': delivery_addr,
                    'equipment': equipment,
                    'service_company': service_comp,
                }
            )
            row_count += 1

        self.stdout.write(self.style.SUCCESS(f"Импорт машин завершён, обработано: {row_count}."))

    # ------------------ 2. Импорт ТО ------------------------
    def import_to(self, filename):
        self.stdout.write(f"Импорт ТО из {filename}...")
        try:
            wb = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f"Файл {filename} не найден!"))
            return

        sheet = wb.active
        row_count = 0


        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 7:
                self.stdout.write("Строка ТО короче 7-8 столбцов, пропускаем: " + str(row))
                continue

            factory_number = row[0] or ''
            to_type = row[1] or ''
            date_performed = self._parse_date(row[2])
            operating_time = row[3] or 0
            order_number = row[4] or ''
            order_date = self._parse_date(row[5])
            organization = row[6] or ''

            service_company = ''
            if len(row) > 7:
                service_company = row[7] or ''

            if not factory_number:
                self.stdout.write("Пустой factory_number (ТО), пропускаем.")
                continue

            try:
                machine = Machine.objects.get(factory_number=factory_number)
            except Machine.DoesNotExist:
                self.stdout.write(f"Машина с Зав.№={factory_number} не найдена, пропускаем.")
                continue

            TO.objects.create(
                to_type=to_type,
                date_performed=date_performed,
                operating_time=operating_time,
                order_number=order_number,
                order_date=order_date,
                organization=organization,
                service_company=service_company,
                machine=machine
            )
            row_count += 1

        self.stdout.write(self.style.SUCCESS(f"Импорт ТО завершён, добавлено: {row_count}."))

    # ------------------ 3. Импорт Рекламаций ------------------------
    def import_claims(self, filename):
        self.stdout.write(f"Импорт рекламаций из {filename}...")
        try:
            wb = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f"Файл {filename} не найден!"))
            return

        sheet = wb.active
        row_count = 0


        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 9:
                self.stdout.write("Строка рекламаций короче 9 столбцов, пропускаем: " + str(row))
                continue

            factory_number = row[0] or ''
            failure_date = self._parse_date(row[1])
            operating_time = row[2] or 0
            failure_node = row[3] or ''
            failure_description = row[4] or ''
            recovery_method = row[5] or ''
            used_parts = row[6] or ''
            recovery_date = self._parse_date(row[7])
            downtime = row[8] or 0

            if not factory_number:
                self.stdout.write("Пустой factory_number (рекламации), пропускаем.")
                continue

            try:
                machine = Machine.objects.get(factory_number=factory_number)
            except Machine.DoesNotExist:
                self.stdout.write(f"Машина с Зав.№={factory_number} не найдена, пропускаем.")
                continue

            Reclamation.objects.create(
                failure_date=failure_date,
                operating_time=operating_time,
                failure_node=failure_node,
                failure_description=failure_description,
                recovery_method=recovery_method,
                used_parts=used_parts,
                recovery_date=recovery_date,
                downtime=downtime,
                # service_company=...,
                machine=machine
            )
            row_count += 1

        self.stdout.write(self.style.SUCCESS(f"Импорт рекламаций завершён, добавлено: {row_count}."))

    # ------------------ Вспомогательная функция для дат ------------------------
    def _parse_date(self, raw_val):
        """
        Пытаемся корректно преобразовать ячейку Excel (datetime или строка)
        в Python date. Если не получилось, вернём None.
        """
        import datetime
        if isinstance(raw_val, datetime.datetime):
            return raw_val.date()
        elif isinstance(raw_val, str):
            try:
                return datetime.datetime.strptime(raw_val, '%Y-%m-%d').date()
            except ValueError:
                pass

            try:
                return datetime.datetime.strptime(raw_val, '%d.%m.%Y').date()
            except ValueError:
                pass

            try:
                return datetime.datetime.strptime(raw_val, '%d/%m/%Y').date()
            except ValueError:
                return None

        return None
